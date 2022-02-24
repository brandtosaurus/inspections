import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from sqlalchemy import create_engine

from win10toast import ToastNotifier

CSV = r"https://web.fulcrumapp.com/shares/48fc49435c5a0199.csv"
WORKBOOK = r"C:\Users\MB2705851\OneDrive - Surbana Jurong Private Limited\PROJECTS\C1579_RRAMS Overberg District Mun\2021\data_do_not_delete.xlsx"
SHEETNAME = "Sheet1"
ROADSHEET = "Sheet2"

ROADS_QRY = """SELECT * FROM infrastructure.asset where district_municipality_id = 50 AND asset_type_id = 2"""

ENGINE = create_engine(
    "postgresql://postgres:$admin@localhost:5432/asset_management_master"
)

toast = ToastNotifier()
toast.show_toast(
    "SCRIPT RUNNING",
    "ODM Progress Document Updating",
    duration=10,
)


df = pd.read_csv(CSV, low_memory=False)
roads_df = pd.read_sql_query(ROADS_QRY, ENGINE)
for col in roads_df.columns:
    try:
        roads_df[col] = roads_df[col].astype(str)
    except:
        pass

roads_df['local_municipality_id'] = roads_df['local_municipality_id'].astype(int)
roads_df['town_id'] = roads_df['town_id'].astype(float)
roads_df['asset_subtype_id'] = roads_df['asset_subtype_id'].astype(float)
roads_df['length'] = roads_df['length'].astype(float)

df.drop(
    [
        "gps_altitude",
        "gps_horizontal_accuracy",
        "gps_vertical_accuracy",
        "gps_speed",
        "gps_course",
        "gps_course",
    ],
    axis=1,
    inplace=True,
)

wb = load_workbook(WORKBOOK, data_only=True)

ws = wb[SHEETNAME]
rows = dataframe_to_rows(df, index=False, header=True)

for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

ws2 = wb[ROADSHEET]
rows2 = dataframe_to_rows(roads_df, index=False, header=True)

for r_idx, row in enumerate(rows2, 1):
    for c_idx, value in enumerate(row, 1):
        ws2.cell(row=r_idx, column=c_idx, value=value)

wb.save(WORKBOOK)

toast.show_toast(
    "SCRIPT RAN SUCCESSFULLY",
    "ODM Progress Document Updated",
    duration=10,
)
